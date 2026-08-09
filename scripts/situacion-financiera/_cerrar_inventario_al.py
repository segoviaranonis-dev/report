# -*- coding: utf-8 -*-
"""Cierra inventario intake AL: cada archivo → rol Sit Fin + totales + estado."""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
PIPE = ROOT / "scripts/situacion-financiera/pipeline"
INTAKE = ROOT / "scripts/situacion-financiera/intake/corte-AL-03-08-26"
OUT = ROOT / "src/lib/situacion-financiera/inventario-intake-al-0308.json"
AUDIT = ROOT / "src/lib/situacion-financiera/audit-mapa-al-0308.json"

sys.path.insert(0, str(PIPE))
from parsers import (  # noqa: E402
    mes_desde_nombre_cheques,
    parse_cheques_vencer,
    parse_pv_prog,
    parse_saldos_detallado,
    parse_saldos_resumen,
)


def _texto(p: Path) -> str:
    try:
        return p.read_text(encoding="cp1252", errors="replace")
    except Exception:
        return p.read_text(encoding="utf-8", errors="replace")


def _sum_gs_en_txt(texto: str) -> tuple[int, int]:
    """Cuenta líneas con importe Gs al final (heurística ERP)."""
    n = 0
    total = 0
    for l in texto.splitlines():
        m = re.search(r"([\d,]{3,})\s+Gs\s*$", l.rstrip())
        if not m:
            continue
        try:
            v = int(m.group(1).replace(",", ""))
        except ValueError:
            continue
        n += 1
        total += v
    return n, total


def _xlsx_sum_numeric(path: Path, sheet: str | None = None, max_rows: int = 400) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    name = sheet if sheet and sheet in wb.sheetnames else wb.sheetnames[0]
    ws = wb[name]
    vals = []
    labels = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i > max_rows:
            break
        cells = list(row)
        for j, c in enumerate(cells):
            if isinstance(c, (int, float)) and abs(c) >= 1000:
                lab = None
                for k in range(j - 1, -1, -1):
                    if isinstance(cells[k], str) and cells[k].strip():
                        lab = cells[k].strip()[:80]
                        break
                vals.append(float(c))
                labels.append(lab)
    wb.close()
    return {
        "sheet": name,
        "n_numeros": len(vals),
        "suma_abs": sum(abs(v) for v in vals),
        "muestra_labels": [x for x in labels if x][:8],
    }


def main():
    items = []

    # --- Cheques TXT ---
    for p in sorted(INTAKE.glob("*CHEQUES*.txt")):
        ym = mes_desde_nombre_cheques(p.name) or "?"
        t = parse_cheques_vencer(p)["totales"]
        items.append(
            {
                "archivo": p.name,
                "tipo": "txt",
                "familia": "cheques",
                "sit_fin": f"CHEQUES A VENCER · {ym}",
                "mol_key": f"cheques:{ym}",
                "origen": "txt",
                "estado": "mapeado_auditado",
                "n": t["n"],
                "gs": t["importe_gs"],
                "nota": "Canon documentado = Σ líneas TXT",
            }
        )

    # --- Saldos ---
    p = INTAKE / "SALDO CLIENTES DETALLADO AL 03-08.txt"
    if p.exists():
        d = parse_saldos_detallado(p)
        tot = d.get("totales", {})
        items.append(
            {
                "archivo": p.name,
                "tipo": "txt",
                "familia": "cxc_detalle",
                "sit_fin": "AGING / SALDO CLIENTES VENCIDOS (OK)",
                "mol_key": "aging:* + clientes:corte",
                "origen": "txt",
                "estado": "mapeado_auditado",
                "n": tot.get("n") or len(d.get("filas", [])),
                "gs": tot.get("saldo_gs") or tot.get("importe_gs"),
                "nota": "Fuente aging 30…>180 y detalle molecular facturas",
            }
        )

    p = INTAKE / "SALDO CLIENTES AL 03-08.txt"
    if p.exists():
        d = parse_saldos_resumen(p)
        tot = d.get("totales", {})
        items.append(
            {
                "archivo": p.name,
                "tipo": "txt",
                "familia": "cxc_resumen",
                "sit_fin": "Control cruzado Σ clientes (no proyecta mes)",
                "mol_key": "clientes:corte",
                "origen": "txt",
                "estado": "mapeado_auditado",
                "n": tot.get("n") or len(d.get("filas", [])),
                "gs": tot.get("saldo_gs") or tot.get("importe_gs"),
                "nota": "Resumen por cliente; Sit Fin mes usa Excel/cuadro Guido",
            }
        )

    # --- PV ---
    p = INTAKE / "PV Y PROG.txt"
    if p.exists():
        d = parse_pv_prog(p)
        tot = d.get("totales", {})
        filas_pv = d.get("filas", [])
        gs_pedido = sum(int(f.get("Importe_Pedido") or 0) for f in filas_pv)
        items.append(
            {
                "archivo": p.name,
                "tipo": "txt",
                "familia": "pv_prog",
                "sit_fin": "PV Y PROG / MERCADERIAS A ENTREGAR (ref)",
                "mol_key": "pv:YYYY-MM",
                "origen": "txt",
                "estado": "mapeado_auditado",
                "n": tot.get("n") or len(filas_pv),
                "gs": gs_pedido,
                "nota": "TXT = universo; fila Sit Fin = verde Guido (subconjunto)",
                "extra": {"por_mes_cuota": tot.get("por_mes_cuota")},
            }
        )

    for name, nota in [
        ("PV Y PROG - EXCEL.xlsx", "Apoyo Excel PV en sistema / no importados"),
        ("PV Y PROG (NO IMPORTADOS).xlsx", "Apoyo Excel PV no importados"),
    ]:
        p = INTAKE / name
        if p.exists():
            info = _xlsx_sum_numeric(p)
            items.append(
                {
                    "archivo": name,
                    "tipo": "xlsx",
                    "familia": "pv_apoyo",
                    "sit_fin": "Apoyo PV (no sustituye TXT)",
                    "mol_key": None,
                    "origen": "excel_apoyo",
                    "estado": "inventariado",
                    "n": info["n_numeros"],
                    "gs": None,
                    "nota": f"Hoja {info['sheet']} · {nota}",
                    "extra": info,
                }
            )

    # --- Gastos ---
    p = INTAKE / "GASTOS OPERATIVOS 26.xlsx"
    if p.exists():
        info = _xlsx_sum_numeric(p, sheet="PREVISION 2026")
        # buscar fila PREVISION / total en sheet
        wb = openpyxl.load_workbook(p, data_only=True)
        ws = wb["PREVISION 2026"] if "PREVISION 2026" in wb.sheetnames else wb.active
        prevision = None
        for row in ws.iter_rows(values_only=True):
            cells = [c for c in row if c is not None]
            joined = " ".join(str(c).upper() for c in cells if isinstance(c, str))
            if "PREVISION" in joined or "TOTAL" in joined:
                nums = [c for c in row if isinstance(c, (int, float))]
                if nums:
                    prevision = float(nums[-1])
        wb.close()
        items.append(
            {
                "archivo": p.name,
                "tipo": "xlsx",
                "familia": "gastos",
                "sit_fin": "PREVISION GASTOS OPERATIVOS (manual)",
                "mol_key": "manual:PREVISION GASTOS OPERATIVOS",
                "origen": "manual",
                "estado": "mapeado_auditado",
                "n": info["n_numeros"],
                "gs": prevision,
                "nota": "Excel apoyo → fila manual Sit Fin",
                "extra": {"sheet": info["sheet"], "prevision_detectada": prevision},
            }
        )

    # --- Bazzar VTO ---
    p = INTAKE / "VTO.BAZZAR AGOSTO26 .xlsx"
    if not p.exists():
        cands = list(INTAKE.glob("VTO.BAZZAR*.xlsx"))
        p = cands[0] if cands else None
    if p and p.exists():
        wb = openpyxl.load_workbook(p, data_only=True)
        sh = (
            "PREVISION PAGOS BAZZAR26"
            if "PREVISION PAGOS BAZZAR26" in wb.sheetnames
            else wb.sheetnames[0]
        )
        ws = wb[sh]
        por_mes = {}
        total_prev = None
        for row in ws.iter_rows(values_only=True):
            fe, prev = row[0] if len(row) > 0 else None, row[1] if len(row) > 1 else None
            if hasattr(fe, "year") and isinstance(prev, (int, float)):
                por_mes[f"{fe.year}-{fe.month:02d}"] = float(prev)
            if fe is None and isinstance(prev, (int, float)):
                total_prev = float(prev)
        wb.close()
        ago = por_mes.get("2026-08")
        items.append(
            {
                "archivo": p.name,
                "tipo": "xlsx",
                "familia": "bazzar",
                "sit_fin": "PAGOS BAZZAR (manual)",
                "mol_key": "bazzar:manual",
                "origen": "manual",
                "estado": "mapeado_auditado",
                "n": len(por_mes),
                "gs": ago or total_prev,
                "nota": "Excel VTO.BAZZAR → fila manual Sit Fin (ago = previsión mes)",
                "extra": {"sheet": sh, "por_mes": por_mes, "total_anio": total_prev},
            }
        )

    # --- Ventas TXT (control / no fila Sit Fin directa) ---
    for name, familia, sit in [
        ("VENTAS CON DTO AGOSTO 26.txt", "ventas_dto", "Control cobros/ventas (fuera fila Sit Fin)"),
        ("VENTAS MENSUALES AGOSTO 26.txt", "ventas_mes", "Control ventas mes (fuera fila Sit Fin)"),
        ("VENTAS POR DIA CONTROL AGOSTO26.txt", "ventas_dia", "Control ventas día (fuera fila Sit Fin)"),
        ("INFORME DE VENTAS BZZ CLI + FACT.txt", "ventas_bzz_cli", "Ventas Bazzar cli+fact (apoyo)"),
        ("INFORME DE VENTAS BZZ FACT + ART.txt", "ventas_bzz_art", "Ventas Bazzar fact+art (apoyo)"),
    ]:
        p = INTAKE / name
        if not p.exists():
            continue
        texto = _texto(p)
        n, gs = _sum_gs_en_txt(texto)
        items.append(
            {
                "archivo": name,
                "tipo": "txt",
                "familia": familia,
                "sit_fin": sit,
                "mol_key": None,
                "origen": "control",
                "estado": "inventariado_control",
                "n": n,
                "gs": gs if gs else None,
                "nota": "No mueve celda Sit Fin; insumo análisis cobros / Bazzar",
            }
        )

    # --- SF AL Excel objetivo ---
    p = INTAKE / "SF AL 03-08.xlsx"
    if p.exists():
        items.append(
            {
                "archivo": p.name,
                "tipo": "xlsx",
                "familia": "sit_fin_objetivo",
                "sit_fin": "Plantilla SIT FIN (todas las filas UI)",
                "mol_key": "excel-al-0308",
                "origen": "excel_objetivo",
                "estado": "mapeado_auditado",
                "n": None,
                "gs": None,
                "nota": "Snapshot UI + filas manuales/cálculo",
            }
        )

    # --- Difícil cobro: inventariado desde Excel Sit Fin (filtro TXT Guido aún no) ---
    excel_path = ROOT / "src/lib/situacion-financiera/excel-al-0308.json"
    dificil_gs = None
    if excel_path.exists():
        for r in json.loads(excel_path.read_text(encoding="utf-8")).get("rows") or []:
            lab = (r.get("label") or "").upper()
            if "TOTAL DIFICIL" in lab.replace("Í", "I"):
                dificil_gs = r.get("gs")
                break
    items.append(
        {
            "archivo": "SF AL 03-08.xlsx (bloque DIF.COBRO)",
            "tipo": "xlsx",
            "familia": "dificil_cobro",
            "sit_fin": "VENCIDOS (DIF.COBRO) / TOTAL DIFICIL",
            "mol_key": "dificil:total",
            "origen": "excel_prevision",
            "estado": "inventariado",
            "n": 12,
            "gs": dificil_gs,
            "nota": "Canon = Excel Guido. TXT sin tipo DIFICIL/SALEMMA — no reusar aging OK",
        }
    )

    por_estado = {}
    for it in items:
        por_estado[it["estado"]] = por_estado.get(it["estado"], 0) + 1

    cobertura = {
        "corte": "AL-03-08-26",
        "n_archivos_intake": len([p for p in INTAKE.iterdir() if p.is_file()]),
        "n_items_mapa": len(items),
        "por_estado": por_estado,
        "completo_intake": all(
            it["estado"]
            in (
                "mapeado_auditado",
                "inventariado",
                "inventariado_control",
                "pendiente_filtro",
            )
            for it in items
        ),
        "items": items,
    }

    # Enriquecer audit existente
    if AUDIT.exists():
        audit = json.loads(AUDIT.read_text(encoding="utf-8"))
    else:
        audit = {}
    audit["inventario_intake"] = {
        **cobertura,
        "pendientes": [it for it in items if it["estado"] == "pendiente_filtro"],
        "control_fuera_sit_fin": [
            it["archivo"] for it in items if it["estado"] == "inventariado_control"
        ],
    }
    AUDIT.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    OUT.write_text(json.dumps(cobertura, ensure_ascii=False, indent=2), encoding="utf-8")

    print("Archivos disco:", cobertura["n_archivos_intake"])
    print("Items mapa:", len(items))
    print("Estados:", por_estado)
    for it in items:
        print(
            f"[{it['estado']:22}] {it['archivo'][:42]:42} → {it['sit_fin'][:50]} gs={it.get('gs')}"
        )
    print("OUT", OUT)


if __name__ == "__main__":
    main()
