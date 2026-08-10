# -*- coding: utf-8 -*-
"""Árbol molecular desde TXT limpios del corte AL — cada Gs con documentación."""
from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
PIPE = ROOT / "scripts/situacion-financiera/pipeline"
INTAKE = ROOT / "scripts/situacion-financiera/intake/corte-AL-03-08-26"
TABLAS_CLIENTES = ROOT / "scripts/situacion-financiera/intake/tablas/clientes.xlsx"
SALDO_TIPO_COBRO = (
    ROOT
    / "scripts/situacion-financiera/intake/saldo-tipo-cobro/07-2026-Saldo_20260701.txt"
)
CADENA_SNAP = (
    ROOT / "scripts/situacion-financiera/intake/tablas/cliente_cadena_snapshot.json"
)
STAGING = (
    ROOT
    / "scripts/situacion-financiera/data/catalogo_local/staging"
    / "0d5c8324-e773-4848-b964-6fc3101446be"
)
OUT = ROOT / "src/lib/situacion-financiera/molecular-al-0308.json"
EXCEL = ROOT / "src/lib/situacion-financiera/excel-al-0308.json"
TASA = 5970.96

sys.path.insert(0, str(PIPE))
from parsers import (  # noqa: E402
    mes_desde_nombre_cheques,
    parse_cheques_vencer,
    parse_saldos_detallado,
)


def usd(gs):
    if gs is None:
        return None
    return round(float(gs) / TASA, 2)


def node(id_, label, gs=None, meta=None, children=None, fuente=None, doc=None):
    n = {"id": id_, "label": label}
    if gs is not None:
        n["gs"] = float(gs)
        n["usd"] = usd(gs)
    if meta:
        n["meta"] = meta
    if fuente:
        n["fuente"] = fuente
    if doc:
        n["doc"] = doc
    if children:
        n["children"] = children
    return n


def top_n(items, key, n=20):
    return sorted(items, key=key, reverse=True)[:n]


def mes_cheque_archivo(nombre: str) -> str:
    return mes_desde_nombre_cheques(nombre) or "s/m"


def mes_from_fecha(fe: str) -> str:
    """Acepta YYYYMMDD, YYYY-MM-DD o dd/mm/yyyy."""
    fe = str(fe or "").strip()
    if not fe:
        return "s/m"
    if "/" in fe:
        parts = fe.split("/")
        if len(parts) == 3:
            return f"{parts[2]}-{parts[1].zfill(2)}"
    digits = "".join(c for c in fe if c.isdigit())
    if len(digits) >= 8:
        return f"{digits[:4]}-{digits[4:6]}"
    if len(fe) >= 7 and fe[4] == "-":
        return fe[:7]
    return "s/m"


def build_cheques_from_txt() -> dict:
    """Mes → día vto → banco → cheque (TXT limpio completo, sin truncar)."""
    out = {}
    files = sorted(INTAKE.glob("*CHEQUES*.txt"))
    by_mes: dict[str, list] = defaultdict(list)
    for path in files:
        parsed = parse_cheques_vencer(path)
        ym = mes_cheque_archivo(path.name)
        for f in parsed["filas"]:
            if f.get("Moneda") != "Gs":
                continue
            by_mes[ym].append(f)

    for mes, rows in sorted(by_mes.items()):
        by_dia: dict[str, list] = defaultdict(list)
        for r in rows:
            by_dia[r.get("Fecha_Vto") or "?"].append(r)
        dia_nodes = []
        for dia, items in sorted(by_dia.items(), key=lambda x: x[0]):
            by_banco: dict[str, list] = defaultdict(list)
            for r in items:
                key = f"{r.get('Banco_Cod')}|{r.get('Banco_Nombre') or ''}"
                by_banco[key].append(r)
            banco_nodes = []
            for bkey, cheqs in sorted(
                by_banco.items(),
                key=lambda x: -sum(int(c["Importe"]) for c in x[1]),
            ):
                cod, nom = bkey.split("|", 1)
                leaves = [
                    node(
                        f"ch-{mes}-{dia}-{cod}-{c['Nro_Cheque']}-{i}",
                        f"Cheque {c['Nro_Cheque']} · {c.get('Emitente') or 's/emitente'} · cli {c.get('Cod_Cliente')}",
                        c["Importe"],
                        meta=f"vto {c.get('Fecha_Vto')} · proc {c.get('Fecha_Proc')} · {c.get('Moneda')} · línea TXT #{c.get('Nro_Linea')}",
                        fuente=c.get("Fuente"),
                        doc=c.get("Linea_Limpia"),
                    )
                    for i, c in enumerate(
                        sorted(cheqs, key=lambda x: -int(x["Importe"]))
                    )
                ]
                banco_nodes.append(
                    node(
                        f"banco-{mes}-{dia}-{cod}",
                        f"{cod} · {nom or 'banco'}",
                        sum(int(c["Importe"]) for c in cheqs),
                        meta=f"{len(cheqs)} cheques · TXT limpio",
                        children=leaves,
                        fuente=cheqs[0].get("Fuente") if cheqs else None,
                    )
                )
            dia_nodes.append(
                node(
                    f"dia-{mes}-{dia}",
                    f"Vence {dia}",
                    sum(int(c["Importe"]) for c in items),
                    meta=f"{len(items)} cheques · subtotal día (como listado ERP)",
                    children=banco_nodes,
                )
            )
        archivos = sorted({r.get("Fuente") for r in rows if r.get("Fuente")})
        out[f"cheques:{mes}"] = node(
            f"cheques-{mes}",
            f"Cheques a vencer {mes}",
            sum(int(c["Importe"]) for c in rows),
            meta=f"{len(rows)} cheques · Σ Gs respaldada 100% por líneas TXT",
            children=dia_nodes,
            fuente=" · ".join(archivos),
        )
    return out


def load_staging(name: str):
    p = STAGING / name
    if not p.exists():
        return []
    return json.loads(p.read_text(encoding="utf-8"))


def build_clientes_facturas() -> dict:
    """Cliente → factura (staging desde TXT saldos)."""
    facturas = load_staging("sf_saldo_factura.json")
    clientes = load_staging("sf_saldo_cliente.json")
    fac_by_cli = defaultdict(list)
    for f in facturas:
        fac_by_cli[str(f.get("cod_cliente") or "")].append(f)

    cli_nodes = []
    for c in top_n(clientes, lambda x: abs(float(x.get("saldo") or 0)), 50):
        cod = str(c.get("cod_cliente") or "")
        facs = sorted(
            fac_by_cli.get(cod, []),
            key=lambda x: abs(float(x.get("saldo") or 0)),
            reverse=True,
        )
        children = [
            node(
                f"fac-{cod}-{i}-{fac.get('nro_factura')}",
                f"Factura {fac.get('nro_factura')}",
                fac.get("saldo"),
                meta=f"días {fac.get('dias_vencido')} · {fac.get('nombre') or ''}",
                fuente=fac.get("archivo"),
                doc=f"factura={fac.get('nro_factura')} cli={cod} saldo={fac.get('saldo')} dias={fac.get('dias_vencido')} · TXT {fac.get('archivo')}",
            )
            for i, fac in enumerate(facs)
        ]
        cli_nodes.append(
            node(
                f"cli-{cod}",
                f"{cod} · {c.get('nombre') or 'sin nombre'}",
                c.get("saldo"),
                meta=str(c.get("moneda") or ""),
                children=children or None,
                fuente=c.get("archivo") or "sf_saldo_cliente",
            )
        )

    root = node(
        "clientes-corte",
        "Saldo clientes (TXT limpio)",
        sum(float(c.get("saldo") or 0) for c in clientes),
        meta=f"{len(clientes)} clientes · {len(facturas)} facturas",
        children=cli_nodes,
        fuente="SALDO CLIENTES*.txt",
    )

    buckets = {
        "v30": (1, 30),
        "v60": (31, 60),
        "v90": (61, 90),
        "v120": (91, 120),
        "v150": (121, 150),
        "v180": (151, 180),
        "v180p": (181, 10_000),
    }
    aging = {}
    for key, (lo, hi) in buckets.items():
        rows = [f for f in facturas if lo <= int(f.get("dias_vencido") or 0) <= hi]
        by_cli = defaultdict(list)
        for f in rows:
            by_cli[str(f.get("cod_cliente"))].append(f)
        children = []
        for cod, facs in top_n(
            list(by_cli.items()),
            lambda x: sum(abs(float(f.get("saldo") or 0)) for f in x[1]),
            40,
        ):
            gs = sum(float(f.get("saldo") or 0) for f in facs)
            children.append(
                node(
                    f"aging-{key}-{cod}",
                    f"Cliente {cod}",
                    gs,
                    children=[
                        node(
                            f"aging-{key}-{cod}-{i}-{fac.get('nro_factura')}",
                            f"Factura {fac.get('nro_factura')}",
                            fac.get("saldo"),
                            meta=f"{fac.get('dias_vencido')} d",
                            fuente=fac.get("archivo"),
                            doc=f"factura={fac.get('nro_factura')} cli={cod} saldo={fac.get('saldo')} · TXT {fac.get('archivo')}",
                        )
                        for i, fac in enumerate(
                            sorted(facs, key=lambda x: -abs(float(x.get("saldo") or 0)))
                        )
                    ],
                )
            )
        aging[f"aging:{key}"] = node(
            f"aging-{key}",
            f"Vencidos {key}",
            sum(float(f.get("saldo") or 0) for f in rows),
            meta=f"{len(rows)} facturas · TXT saldos detallado",
            children=children,
            fuente="SALDO CLIENTES DETALLADO*.txt",
        )

    index = {"clientes:corte": root, **aging}
    for mes in ["2026-08", "2026-09", "2026-10", "2026-11", "2026-12"]:
        index[f"clientes:{mes}"] = node(
            f"clientes-{mes}",
            f"Saldo clientes {mes}",
            root.get("gs"),
            meta="Proxy corte (explosión cuotas Guido pendiente) · mismas facturas TXT",
            children=cli_nodes,
            fuente="SALDO CLIENTES*.txt",
        )
    return index


def build_pv() -> dict:
    pv = load_staging("sf_pv_prog.json")
    by_mes = defaultdict(list)
    for p in pv:
        by_mes[mes_from_fecha(p.get("Fecha_Entrega"))].append(p)
    out = {}
    for mes, rows in sorted(by_mes.items()):
        by_cli = defaultdict(list)
        for p in rows:
            by_cli[str(p.get("Cod_Cliente") or "?")].append(p)
        cli_kids = []
        for cod, items in sorted(
            by_cli.items(),
            key=lambda x: -sum(
                float(i.get("Importe_Pedido") or i.get("Importe_Cuota") or 0) for i in x[1]
            ),
        ):
            kids = [
                node(
                    f"pv-{mes}-{cod}-{i}-{p.get('Nro_Ped_Prov')}",
                    f"Ped.prov {p.get('Nro_Ped_Prov')} · PF {p.get('Proforma')}",
                    p.get("Importe_Pedido") or p.get("Importe_Cuota"),
                    meta=f"entrega {p.get('Fecha_Entrega')} · cuotas {p.get('Cant_Cuotas')}",
                    fuente=p.get("Fuente") or "PV Y PROG.txt",
                    doc=f"ped={p.get('Nro_Ped_Prov')} pf={p.get('Proforma')} cli={cod} importe={p.get('Importe_Pedido') or p.get('Importe_Cuota')} · TXT PV Y PROG",
                )
                for i, p in enumerate(
                    sorted(
                        items,
                        key=lambda x: -abs(
                            float(x.get("Importe_Pedido") or x.get("Importe_Cuota") or 0)
                        ),
                    )
                )
            ]
            tot = sum(float(i.get("Importe_Pedido") or i.get("Importe_Cuota") or 0) for i in items)
            cli_kids.append(
                node(f"pv-cli-{mes}-{cod}", f"Cliente {cod}", tot, meta=f"{len(items)} líneas", children=kids)
            )
        out[f"pv:{mes}"] = node(
            f"pv-{mes}",
            f"PV y PROG {mes}",
            sum(float(i.get("Importe_Pedido") or i.get("Importe_Cuota") or 0) for i in rows),
            meta=f"{len(rows)} líneas · TXT limpio",
            children=cli_kids,
            fuente="PV Y PROG.txt",
        )
        out[f"mercaderia:{mes}"] = out[f"pv:{mes}"]
    return out


def build_bancos_manual() -> dict:
    # Bancos = MANUAL Excel (no vienen del TXT cheques). Documentar explícito.
    specs = [
        ("banco:CONTINENTAL:USD", "Banco Continental USD", 925101193.7736),
        ("banco:CONTINENTAL:GS", "Banco Continental Gs", 1718236201.0),
        ("banco:ITAU:GS", "Banco Itaú Gs", 359782579.0),
        ("banco:BANCOOP:USD", "Bancoop USD", 2352986596.6704),
        ("banco:BANCOOP:GS", "Bancoop Gs", 990211535.0),
        ("banco:GNB:GS", "GNB Gs", 10000000.0),
        ("banco:BNF:GS", "BNF Gs", 15006624.0),
    ]
    out = {}
    for k, lab, gs in specs:
        out[k] = node(
            k.replace(":", "-"),
            lab,
            gs,
            meta="MANUAL Excel — no hay TXT extracto bancario en intake; no desglose molecular ERP",
            children=[
                node(
                    f"{k}-leaf",
                    "Importe planilla SF AL (sin respaldo TXT)",
                    gs,
                    doc="Sin Linea_Limpia: saldo de banco se carga a mano en Excel.",
                )
            ],
            fuente="SF AL 03-08.xlsx · MANUAL",
        )
    egresos = {
        "manual:PAGO A PROVEEDORES": (-280266114.672, "PAGO A PROVEEDORES"),
        "manual:GASTOS DE DESPACHO": (-115000000.0, "GASTOS DE DESPACHO"),
        "manual:PREVISION GASTOS OPERATIVOS": (-2338191000.0, "PREVISION GASTOS OPERATIVOS"),
        "manual:PRESTAMO BANCARIO": (-157000000.0, "PRESTAMO BANCARIO"),
    }
    for k, (gs, lab) in egresos.items():
        out[k] = node(
            k.replace(":", "-"),
            lab,
            gs,
            meta="MANUAL Excel — respaldo = planilla / comprobante, no TXT listado ERP",
            children=[
                node(f"{k}-leaf", "Importe planilla", gs, doc=f"Manual Excel · {lab}")
            ],
            fuente="SF AL · MANUAL",
        )
    out["bazzar:manual"] = node(
        "bazzar",
        "Pagos Bazzar",
        1_300_000_000.0,
        meta="MANUAL · VTO.BAZZAR previsión ago-26",
        children=[
            node(
                "bazzar-ago",
                "Previsión agosto 2026",
                1_300_000_000.0,
                doc="VTO.BAZZAR AGOSTO26.xlsx · hoja PREVISION PAGOS BAZZAR26 · 2026-08 = 1.300.000.000",
            )
        ],
        fuente="VTO.BAZZAR AGOSTO26.xlsx",
    )
    return out


def load_tipos_cobro(path: Path | None = None) -> dict[str, str]:
    """CODIGO → TIPO COBRO (col C): OK, LUISITO, DIFICIL, SALEMMA, …"""
    import openpyxl

    p = path or TABLAS_CLIENTES
    wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    out: dict[str, str] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1 or not row:
            continue
        cod, tipo = row[0], row[2] if len(row) > 2 else None
        if cod is None or tipo is None:
            continue
        if isinstance(cod, (int, float)):
            cod_k = str(int(cod))
        else:
            cod_k = str(cod).strip()
            if cod_k.isdigit():
                cod_k = str(int(cod_k))
        out[cod_k] = str(tipo).strip().upper()
    wb.close()
    return out


def _aging_bucket(dias: int) -> str:
    if dias < 0:
        return "no_vencido"
    if dias <= 30:
        return "v30"
    if dias <= 60:
        return "v60"
    if dias <= 90:
        return "v90"
    if dias <= 120:
        return "v120"
    if dias <= 150:
        return "v150"
    if dias <= 180:
        return "v180"
    return "v180p"


def load_cadena_map(path: Path | None = None) -> dict[str, dict]:
    """id_cliente → primaria cadena (snapshot Nexus cliente_cadena_v2)."""
    p = path or CADENA_SNAP
    if not p.exists():
        alt = ROOT / "src/lib/situacion-financiera/cliente-cadena-snapshot.json"
        p = alt if alt.exists() else p
    if not p.exists():
        return {}
    data = json.loads(p.read_text(encoding="utf-8"))
    out = {}
    for k, v in (data.get("por_id") or {}).items():
        out[str(k).strip()] = v
    return out


def _cadena_de(cod: str, cadena_map: dict[str, dict]) -> tuple[str, str]:
    """(id_cadena_key, label)."""
    info = cadena_map.get(str(cod).strip()) or {}
    prim = info.get("primaria") or {}
    if prim.get("id_cadena") is not None:
        cid = str(prim["id_cadena"])
        nom = (prim.get("descp_cadena") or f"Cadena {cid}").strip()
        return cid, nom
    return "sin", "SIN CADENA (cliente_cadena_v2)"


def _tree_clientes_facturas(
    filas: list,
    *,
    id_prefix: str,
    label: str,
    meta: str,
    fuente: str,
    cadena_map: dict[str, dict] | None = None,
) -> dict:
    """Árbol: [Cadena →] Cliente → Factura (Linea_Limpia)."""
    cmap = cadena_map if cadena_map is not None else load_cadena_map()
    by_cli: dict[str, list] = defaultdict(list)
    for f in filas:
        by_cli[str(f.get("Cod_Cliente") or "?")].append(f)

    # Agrupar clientes por cadena
    by_cad: dict[str, dict] = {}
    for cod, facts in by_cli.items():
        cid, cnom = _cadena_de(cod, cmap)
        if cid not in by_cad:
            by_cad[cid] = {"nombre": cnom, "clientes": {}}
        by_cad[cid]["clientes"][cod] = facts

    cadena_nodes = []
    total = 0.0
    for cid, pack in sorted(
        by_cad.items(),
        key=lambda x: -sum(
            int(i.get("Saldo") or 0)
            for facts in x[1]["clientes"].values()
            for i in facts
        ),
    ):
        cli_nodes = []
        cad_gs = 0.0
        for cod, facts in sorted(
            pack["clientes"].items(),
            key=lambda x: -sum(int(i.get("Saldo") or 0) for i in x[1]),
        ):
            gs = float(sum(int(i.get("Saldo") or 0) for i in facts))
            cad_gs += gs
            nom = (facts[0].get("Nombre") or "").strip() or f"Cliente {cod}"
            info = cmap.get(str(cod).strip()) or {}
            tipo_cli = info.get("tipo_cliente") or ""
            fac_nodes = []
            for i, fac in enumerate(
                sorted(facts, key=lambda x: -int(x.get("Saldo") or 0))
            ):
                s = float(fac.get("Saldo") or 0)
                fac_nodes.append(
                    node(
                        f"{id_prefix}-f-{cod}-{i}",
                        f"{fac.get('Nro_Factura')} · {fac.get('Dias_Vencido')}d",
                        s,
                        doc=fac.get("Linea_Limpia")
                        or f"Factura {fac.get('Nro_Factura')} saldo {s}",
                        fuente=fuente,
                    )
                )
            cli_nodes.append(
                node(
                    f"{id_prefix}-c-{cod}",
                    f"{nom} ({cod})",
                    gs,
                    meta=(
                        f"{len(facts)} factura(s)"
                        + (f" · {tipo_cli}" if tipo_cli else "")
                        + f" · cadena={pack['nombre']}"
                    ),
                    children=fac_nodes,
                    fuente=fuente,
                )
            )
        total += cad_gs
        cadena_nodes.append(
            node(
                f"{id_prefix}-cad-{cid}",
                f"Cadena · {pack['nombre']}",
                cad_gs,
                meta=f"{len(cli_nodes)} cliente(s) · id_cadena={cid}",
                children=cli_nodes,
                fuente="cliente_cadena_v2 (tabla maestra) + " + fuente,
            )
        )

    # Si solo hay una cadena, aplanar un nivel menos molesta? No: siempre mostrar
    # cadena para auditar potencial holding.
    return node(
        id_prefix,
        label,
        total,
        meta=meta + f" · {len(cadena_nodes)} cadena(s)",
        children=cadena_nodes,
        fuente=fuente,
    )


def build_tipo_cobro_saldo_txt() -> dict:
    """Cruce clientes.xlsx (TIPO COBRO) × saldo TXT × cliente_cadena_v2."""
    if not TABLAS_CLIENTES.exists() or not SALDO_TIPO_COBRO.exists():
        return {}
    tipos = load_tipos_cobro()
    cadena_map = load_cadena_map()
    parsed = parse_saldos_detallado(SALDO_TIPO_COBRO)
    fuente = SALDO_TIPO_COBRO.name
    by_tipo: dict[str, list] = defaultdict(list)
    sin_tipo = 0
    for f in parsed["filas"]:
        cod = str(f.get("Cod_Cliente") or "").strip()
        t = tipos.get(cod)
        if not t:
            sin_tipo += 1
            continue
        f = dict(f)
        f["Tipo_Cobro"] = t
        by_tipo[t].append(f)

    out: dict = {}
    # LUISITO → acordeón PAGO LUISITO
    luisito_filas = by_tipo.get("LUISITO") or []
    if luisito_filas:
        tree = _tree_clientes_facturas(
            luisito_filas,
            id_prefix="luisito-txt",
            label="PAGO LUISITO · saldo TXT × tipo cobro × cadena",
            meta=(
                f"TXT {fuente} · clientes.xlsx TIPO=LUISITO · "
                f"cliente_cadena_v2 · "
                f"{len({x['Cod_Cliente'] for x in luisito_filas})} clientes · "
                f"{len(luisito_filas)} facturas · Σ Gs respaldada al peso"
            ),
            fuente=f"clientes.xlsx + maestra cliente_cadena_v2 + {fuente}",
            cadena_map=cadena_map,
        )
        out["luisito:cuadro"] = tree
        for mes in ["2026-08", "2026-09", "2026-10", "2026-11", "2026-12", "2026-07"]:
            out[f"luisito:{mes}"] = tree

    # DIFICIL + SALEMMA → DIF.COBRO (misma regla Guido)
    dificil_filas = (by_tipo.get("DIFICIL") or []) + (by_tipo.get("SALEMMA") or [])
    if dificil_filas:
        tree_d = _tree_clientes_facturas(
            dificil_filas,
            id_prefix="dificil-txt",
            label="DIF.COBRO · DIFICIL+SALEMMA (TXT × cadena maestra)",
            meta=(
                f"TXT {fuente} · tipos DIFICIL+SALEMMA · cliente_cadena_v2 maestra · "
                f"{len({x['Cod_Cliente'] for x in dificil_filas})} clientes · "
                f"{len(dificil_filas)} facturas"
            ),
            fuente=f"clientes.xlsx + maestra cliente_cadena_v2 + {fuente}",
            cadena_map=cadena_map,
        )
        out["dificil:total"] = tree_d
        # aging buckets solo universo difícil
        buckets: dict[str, list] = defaultdict(list)
        for f in dificil_filas:
            buckets[_aging_bucket(int(f.get("Dias_Vencido") or 0))].append(f)
        for b, key in [
            ("v30", "dificil:v30"),
            ("v60", "dificil:v60"),
            ("v90", "dificil:v90"),
            ("v120", "dificil:v120"),
            ("v150", "dificil:v150"),
            ("v180", "dificil:v180"),
            ("v180p", "dificil:v180p"),
        ]:
            if not buckets.get(b):
                continue
            out[key] = _tree_clientes_facturas(
                buckets[b],
                id_prefix=f"dificil-{b}",
                label=f"DIF.COBRO {b} · TXT × cadena",
                meta=f"Filtro tipo DIFICIL/SALEMMA · bucket {b}",
                fuente=f"clientes.xlsx + maestra cliente_cadena_v2 + {fuente}",
                cadena_map=cadena_map,
            )

    # Control OK (no pisa Sit Fin mes; clave de auditoría)
    ok_filas = by_tipo.get("OK") or []
    if ok_filas:
        out["tipo_cobro:OK"] = _tree_clientes_facturas(
            ok_filas,
            id_prefix="ok-txt",
            label="Clientes OK · saldo TXT × cadena maestra",
            meta=f"Control · {len(ok_filas)} facturas · no sustituye proyección mes Excel",
            fuente=f"clientes.xlsx + maestra cliente_cadena_v2 + {fuente}",
            cadena_map=cadena_map,
        )

    print(
        "tipo_cobro",
        {k: len(v) for k, v in sorted(by_tipo.items())},
        "sin_tipo",
        sin_tipo,
        "cadena_map",
        len(cadena_map),
        "luisito_gs",
        out.get("luisito:cuadro", {}).get("gs"),
        "dificil_gs",
        out.get("dificil:total", {}).get("gs"),
    )
    return out


def build_luisito_excel() -> dict:
    """Fallback Excel AL si aún no hay cruce TXT (no debería usarse si hay intake)."""
    if TABLAS_CLIENTES.exists() and SALDO_TIPO_COBRO.exists():
        return {}
    excel = json.loads(EXCEL.read_text(encoding="utf-8"))
    out = {}
    mes_ctx = "2026-08"
    for r in excel.get("rows") or []:
        if r.get("mes"):
            mes_ctx = r["mes"]
        lab = (r.get("label") or "").upper()
        if "LUISITO" not in lab:
            continue
        if r.get("kind") not in ("row",):
            continue
        gs = r.get("gs")
        key = f"luisito:{mes_ctx}"
        out[key] = node(
            key.replace(":", "-"),
            r.get("label") or "PAGO LUISITO",
            float(gs) if gs is not None else 0.0,
            meta="Excel Sit Fin · sin clientes.xlsx/TXT",
            children=[
                node(
                    f"{key}-excel",
                    "Importe celda Excel",
                    float(gs) if gs is not None else 0.0,
                    doc=f"SF AL · fila {r.get('r')} · mes {mes_ctx}",
                )
            ],
            fuente="SF AL 03-08.xlsx",
        )
    return out


def _dificil_key_from_label(label: str) -> str | None:
    u = " ".join((label or "").upper().split())
    if "DIF" not in u or "COBRO" not in u:
        return None
    if "TOTAL" in u:
        return "dificil:total"
    if "MAYOR" in u and "180" in u:
        return "dificil:v180p"
    for n, k in [
        (30, "dificil:v30"),
        (60, "dificil:v60"),
        (90, "dificil:v90"),
        (120, "dificil:v120"),
        (150, "dificil:v150"),
        (180, "dificil:v180"),
    ]:
        if f"VENCIDOS A {n}" in u or f"{n} DIA" in u:
            return k
    # Mes embebido tipo AGOSTO 26 / SETIEMBRE 26
    m = re.search(
        r"\b(ENERO|FEBRERO|MARZO|ABRIL|MAYO|JUNIO|JULIO|AGOSTO|SEPTIEMBRE|SETIEMBRE|OCTUBRE|NOVIEMBRE|DICIEMBRE)\s*(\d{2,4})\b",
        u,
    )
    if not m:
        return "dificil:total"
    mes_es = {
        "ENERO": "01",
        "FEBRERO": "02",
        "MARZO": "03",
        "ABRIL": "04",
        "MAYO": "05",
        "JUNIO": "06",
        "JULIO": "07",
        "AGOSTO": "08",
        "SEPTIEMBRE": "09",
        "SETIEMBRE": "09",
        "OCTUBRE": "10",
        "NOVIEMBRE": "11",
        "DICIEMBRE": "12",
    }
    yy = m.group(2)
    if len(yy) == 2:
        yy = "20" + yy
    return f"dificil:{yy}-{mes_es[m.group(1)]}"


def build_dificil_excel() -> dict:
    """Buckets DIF.COBRO mes/Excel; aging/total se pisan con TXT en build_tipo_cobro."""
    excel = json.loads(EXCEL.read_text(encoding="utf-8"))
    out = {}
    for r in excel.get("rows") or []:
        if r.get("kind") not in ("row", "total_gray"):
            continue
        lab = r.get("label") or ""
        key = _dificil_key_from_label(lab)
        if not key:
            continue
        # aging/total vienen del cruce TXT — no pisar con Excel aquí
        if key.startswith("dificil:v") or key == "dificil:total":
            continue
        gs = r.get("gs")
        if gs is None:
            continue
        out[key] = node(
            key.replace(":", "-"),
            lab,
            float(gs),
            meta="Excel Sit Fin · DIF.COBRO mes (proyección Guido)",
            children=[
                node(
                    f"{key}-excel",
                    "Importe celda Excel",
                    float(gs),
                    doc=f"SF AL 03-08.xlsx · fila {r.get('r')} · {lab}",
                )
            ],
            fuente="SF AL 03-08.xlsx",
        )
    return out


def main():
    index = {}
    index.update(build_cheques_from_txt())
    index.update(build_clientes_facturas())
    index.update(build_pv())
    index.update(build_bancos_manual())
    index.update(build_dificil_excel())
    index.update(build_tipo_cobro_saldo_txt())
    index.update(build_luisito_excel())
    for mes in ["2026-08", "2026-09", "2026-10", "2026-11", "2026-12"]:
        index.setdefault(
            f"cheques:{mes}",
            node(f"cheques-empty-{mes}", f"Cheques {mes}", 0, meta="sin TXT en intake"),
        )
        index.setdefault(
            f"pv:{mes}",
            node(f"pv-empty-{mes}", f"PV {mes}", 0, meta="sin filas"),
        )
        index.setdefault(f"mercaderia:{mes}", index[f"pv:{mes}"])

    OUT.write_text(json.dumps(index, ensure_ascii=False), encoding="utf-8")
    ch = index.get("cheques:2026-08")
    print(
        "ok",
        OUT,
        "keys",
        len(index),
        "bytes",
        OUT.stat().st_size,
        "cheques_ago_n",
        ch.get("meta") if ch else None,
        "gs",
        ch.get("gs") if ch else None,
    )


if __name__ == "__main__":
    main()
