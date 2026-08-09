# -*- coding: utf-8 -*-
"""Genera Excel Situación Financiera (hoja SIT FIN) a partir de agregados TXT + manuales.

Mejora vs Excel del funcionario:
  - Linaje: cada celda auto trae nota de fuente/archivo
  - Bloques mensuales consistentes
  - USD = GS / tasa (explícita)
  - CSV satélite con desglose cheques / aging / PV
"""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.comments import Comment


MESES_ES = {
    1: "ENERO",
    2: "FEBRERO",
    3: "MARZO",
    4: "ABRIL",
    5: "MAYO",
    6: "JUNIO",
    7: "JULIO",
    8: "AGOSTO",
    9: "SEPTIEMBRE",
    10: "OCTUBRE",
    11: "NOVIEMBRE",
    12: "DICIEMBRE",
}


def _money_font():
    return Font(name="Calibri", size=11)


def generar_sit_fin(
    *,
    corte: date,
    tasa_usd: float,
    bancos: list[dict[str, Any]],
    cheques_por_mes: dict[str, int],
    saldo_clientes_por_mes: dict[str, int],
    pv_prog_por_mes: dict[str, int],
    aging: dict[str, int],
    manuales: dict[str, Any],
    dest: Path,
    meta: dict[str, Any] | None = None,
) -> Path:
    """
    bancos: [{label, gs?, usd?}]  — si trae usd, gs = usd*tasa; si trae gs, usd=gs/tasa
    manuales: previsiones por mes y conceptos (gastos, bazzar, luisito, proveedores…)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "SIT FIN"

    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    header_fill = PatternFill("solid", fgColor="0F3D3E")
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor="E8F3F1")
    total_fill = PatternFill("solid", fgColor="FFF3CD")
    auto_fill = PatternFill("solid", fgColor="E7F5FF")  # derivado de TXT
    man_fill = PatternFill("solid", fgColor="F8F9FA")

    ws["C1"] = "TASA DE CAMBIO DÓLAR A LA FECHA"
    ws["D1"] = tasa_usd
    ws["C2"] = f"PREVISION A LA FECHA {corte.strftime('%d/%m/%Y')}"
    ws["D2"] = ws["C2"].value
    ws["C3"] = "SALDOS"
    ws["D3"] = "IMPORTE EN GS"
    ws["E3"] = "IMPORTE EN USD"
    for col in ("C", "D", "E"):
        ws[f"{col}3"].fill = header_fill
        ws[f"{col}3"].font = header_font

    row = 4
    bank_start = row
    for b in bancos:
        label = b["label"]
        if b.get("usd") is not None and b.get("gs") is None:
            usd = float(b["usd"])
            gs = usd * tasa_usd
        else:
            gs = float(b.get("gs") or 0)
            usd = gs / tasa_usd if tasa_usd else 0
        ws[f"C{row}"] = label
        ws[f"D{row}"] = gs
        ws[f"E{row}"] = usd
        ws[f"D{row}"].number_format = "#,##0"
        ws[f"E{row}"].number_format = "#,##0.00"
        ws[f"C{row}"].fill = man_fill
        ws[f"D{row}"].fill = man_fill
        comment = Comment("Entrada manual (bancos) — no viene de TXT ERP", "Nexus SF")
        ws[f"C{row}"].comment = comment
        row += 1
    bank_end = row - 1

    # Bloque mes corriente (resto del mes de corte) — cheques/saldos auto
    ym0 = f"{corte.year}-{corte.month:02d}"

    def put_line(r, label, gs, fuente: str, auto: bool):
        ws[f"C{r}"] = label
        ws[f"D{r}"] = gs if gs is not None else 0
        ws[f"E{r}"] = (gs or 0) / tasa_usd if tasa_usd else 0
        ws[f"D{r}"].number_format = "#,##0"
        ws[f"E{r}"].number_format = "#,##0.00"
        fill = auto_fill if auto else man_fill
        ws[f"C{r}"].fill = fill
        ws[f"D{r}"].fill = fill
        ws[f"E{r}"].fill = fill
        if fuente:
            ws[f"C{r}"].comment = Comment(fuente, "Nexus SF")

    # Agosto (mes corte): líneas operativas
    conceptos_mes = [
        ("CHEQUES A VENCER", "cheques", True),
        ("SALDO DE CLIENTES", "saldo_clientes", True),
        ("MERCADERIAS A ENTREGAR", "mercaderias", False),
        ("SALDO DE CLIENTES VENCIDOS A 30 DIAS", "v30", True),
        ("SALDO DE CLIENTES VENCIDOS A 60 DIAS", "v60", True),
        ("PAGOS BAZZAR", "bazzar", False),
        ("PV Y PROG A COBRAR", "pv_prog", True),
        ("PAGO LUISITO", "luisito", False),
        ("PAGO A PROVEEDORES", "proveedores", False),
        ("GASTOS DE DESPACHO", "despacho", False),
        ("PREVISION GASTOS OPERATIVOS", "gastos_op", False),
        ("PRESTAMO BANCARIO", "prestamo", False),
    ]

    man_mes = manuales.get("meses", {}).get(ym0, {})
    put_line(
        row,
        "CHEQUES A VENCER",
        cheques_por_mes.get(ym0, man_mes.get("cheques", 0)),
        f"AUTO TXT cheques_vencer mes {ym0}",
        True,
    )
    row += 1
    put_line(
        row,
        "SALDO DE CLIENTES",
        # en mes corte el Excel deja vacío a veces; usamos total aging no_vencido+ parcial
        man_mes.get("saldo_clientes", saldo_clientes_por_mes.get(ym0, 0)),
        f"AUTO/manual saldo clientes {ym0}",
        ym0 in saldo_clientes_por_mes,
    )
    row += 1
    put_line(row, "MERCADERIAS A ENTREGAR", man_mes.get("mercaderias", 0), "Manual / PV no facturado", False)
    row += 1
    put_line(row, "SALDO DE CLIENTES VENCIDOS A 30 DIAS", aging.get("v30", 0), "AUTO saldos_detallado aging", True)
    row += 1
    put_line(row, "SALDO DE CLIENTES VENCIDOS A 60 DIAS", aging.get("v60", 0), "AUTO saldos_detallado aging", True)
    row += 1
    put_line(row, "PAGOS BAZZAR", man_mes.get("bazzar", 0), "Manual / VTO.BAZZAR", False)
    row += 1
    put_line(
        row,
        "PV Y PROG A COBRAR",
        pv_prog_por_mes.get(ym0, man_mes.get("pv_prog", 0)),
        f"AUTO PV Y PROG.txt cuotas {ym0}",
        True,
    )
    row += 1
    for key, label in [
        ("luisito", "PAGO LUISITO"),
        ("proveedores", "PAGO A PROVEEDORES"),
        ("despacho", "GASTOS DE DESPACHO"),
        ("gastos_op", "PREVISION GASTOS OPERATIVOS"),
        ("prestamo", "PRESTAMO BANCARIO"),
    ]:
        val = man_mes.get(key, 0)
        put_line(row, label, val, "Manual (previsión)", False)
        row += 1

    total_row = row
    ws[f"C{total_row}"] = f"SALDO DISPONIBLE {MESES_ES[corte.month]} {corte.year}"
    # suma desde bancos + líneas del mes
    first_op = bank_end + 1
    ws[f"D{total_row}"] = f"=SUM(D{bank_start}:D{total_row-1})"
    ws[f"E{total_row}"] = f"=SUM(E{bank_start}:E{total_row-1})"
    ws[f"C{total_row}"].fill = total_fill
    ws[f"D{total_row}"].fill = total_fill
    ws[f"E{total_row}"].fill = total_fill
    ws[f"D{total_row}"].number_format = "#,##0"
    ws[f"E{total_row}"].number_format = "#,##0.00"
    row += 1

    # Meses siguientes: sep..dic del mismo año + opcional
    prev_total = total_row
    for month in range(corte.month + 1, 13):
        ym = f"{corte.year}-{month:02d}"
        mm = manuales.get("meses", {}).get(ym, {})
        ws[f"B{row}"] = date(corte.year, month, 1)
        put_line(
            row,
            "CHEQUES A VENCER",
            cheques_por_mes.get(ym, mm.get("cheques", 0)),
            f"AUTO TXT cheques {ym}",
            ym in cheques_por_mes,
        )
        row += 1
        put_line(
            row,
            "SALDO DE CLIENTES",
            saldo_clientes_por_mes.get(ym, mm.get("saldo_clientes", 0)),
            f"Saldo clientes bucket {ym} (manual si no hay proyección mensual)",
            ym in saldo_clientes_por_mes,
        )
        row += 1
        for key, label, auto_key in [
            ("mercaderias", "MERCADERIAS A ENTREGAR", None),
            ("bazzar", "PAGOS DE BAZZAR", None),
            ("luisito", "PAGO LUISITO", None),
            ("pv_prog", "PV Y PROG A COBRAR", "pv"),
            ("proveedores", "PAGO A PROVEEDORES", None),
            ("despacho", "GASTOS DE DESPACHO", None),
            ("gastos_op", "PREVISION GASTOS OPERATIVOS", None),
            ("prestamo", "PRESTAMO BANCARIO", None),
        ]:
            if key == "pv_prog":
                val = pv_prog_por_mes.get(ym, mm.get("pv_prog", 0))
                auto = ym in pv_prog_por_mes
                src = f"AUTO PV Y PROG {ym}" if auto else "Manual"
            else:
                val = mm.get(key, 0)
                auto = False
                src = "Manual (previsión)"
            put_line(row, label, val, src, auto)
            row += 1
        ws[f"C{row}"] = f"SALDO DISPONIBLE {MESES_ES[month]} {corte.year}"
        ws[f"D{row}"] = f"=SUM(D{prev_total}:D{row-1})"
        ws[f"E{row}"] = f"=SUM(E{prev_total}:E{row-1})"
        for col in ("C", "D", "E"):
            ws[f"{col}{row}"].fill = total_fill
        ws[f"D{row}"].number_format = "#,##0"
        ws[f"E{row}"].number_format = "#,##0.00"
        prev_total = row
        row += 2

    # Sección aging / difícil cobro (auto)
    row += 1
    ws[f"C{row}"] = "SALDO DE CLIENTES A COBRAR (AGING AUTO)"
    ws[f"C{row}"].fill = section_fill
    ws[f"C{row}"].font = Font(bold=True)
    row += 1
    for key, label in [
        ("v30", "VENCIDOS A 30 DIAS"),
        ("v60", "VENCIDOS A 60 DIAS"),
        ("v90", "VENCIDOS A 90 DIAS"),
        ("v120", "VENCIDOS A 120 DIAS"),
        ("v150", "VENCIDOS A 150 DIAS"),
        ("v180", "VENCIDOS A 180 DIAS"),
        ("v180p", "VENCIDOS MAYOR A 180 DIAS"),
        ("no_vencido", "NO VENCIDOS (D.VDOS < 0)"),
    ]:
        put_line(row, label, aging.get(key, 0), "AUTO saldos_detallado", True)
        row += 1

    # Meta
    row += 2
    ws[f"C{row}"] = "META NEXUS"
    ws[f"C{row}"].font = Font(bold=True, color="0F3D3E")
    row += 1
    ws[f"C{row}"] = "Azul claro = derivado de TXT · Gris = manual · Dorado = total"
    row += 1
    if meta:
        ws[f"C{row}"] = f"Corte procesado: {meta.get('corte_id', '')}"
        row += 1
        ws[f"C{row}"] = f"Archivos clasificados: {meta.get('n_archivos', '')}"

    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 16

    # hoja LINAGE
    w2 = wb.create_sheet("LINAJE")
    w2["A1"] = "Concepto"
    w2["B1"] = "Valor GS"
    w2["C1"] = "Fuente"
    w2["D1"] = "Tipo"
    r = 2
    for ym, v in sorted(cheques_por_mes.items()):
        w2[f"A{r}"] = f"CHEQUES {ym}"
        w2[f"B{r}"] = v
        w2[f"C{r}"] = "cheques_vencer TXT"
        w2[f"D{r}"] = "AUTO"
        r += 1
    for ym, v in sorted(pv_prog_por_mes.items()):
        w2[f"A{r}"] = f"PV_PROG {ym}"
        w2[f"B{r}"] = v
        w2[f"C{r}"] = "PV Y PROG.txt"
        w2[f"D{r}"] = "AUTO"
        r += 1
    for k, v in sorted(aging.items()):
        w2[f"A{r}"] = f"AGING {k}"
        w2[f"B{r}"] = v
        w2[f"C{r}"] = "saldos_detallado"
        w2[f"D{r}"] = "AUTO"
        r += 1

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    return dest
