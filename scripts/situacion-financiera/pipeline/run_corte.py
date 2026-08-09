# -*- coding: utf-8 -*-
"""
Orquestador corte Situación Financiera AL.
Uso:
  python run_corte.py --entrada PATH_INTAKE --salida PATH_OUT
  python run_corte.py   # defaults intake/corte-AL-03-08-26
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

from clasificador import clasificar_carpeta  # noqa: E402
from parsers import (  # noqa: E402
    mes_desde_nombre_cheques,
    parse_por_tipo,
    escribir_csv,
)
from generar_sit_fin import generar_sit_fin  # noqa: E402
from persistencia import (  # noqa: E402
    CatalogoLocal,
    persistir_corte_local,
    persistir_supabase,
)

try:
    import openpyxl
except ImportError:
    openpyxl = None


DEFAULT_IN = (
    Path(__file__).resolve().parents[1] / "intake" / "corte-AL-03-08-26"
)
DEFAULT_OUT = Path(__file__).resolve().parents[1] / "out" / "AL-03-08-26"


def extraer_manuales_referencia(xlsx: Path) -> dict:
    """Lee SF AL referencia: bancos + líneas manuales por mes (valores literales)."""
    if openpyxl is None or not xlsx.exists():
        return {"tasa_usd": 5970.96, "bancos": [], "meses": {}}
    wb = openpyxl.load_workbook(xlsx, data_only=False)
    ws = wb["SIT FIN"] if "SIT FIN" in wb.sheetnames else wb.active
    tasa = ws["D1"].value or 5970.96
    if isinstance(tasa, str):
        try:
            tasa = float(tasa)
        except ValueError:
            tasa = 5970.96

    bancos = []
    for r in range(4, 11):
        label = ws.cell(r, 3).value
        if not label:
            continue
        d = ws.cell(r, 4).value
        e = ws.cell(r, 5).value
        # prefer numeric literals
        gs = None
        usd = None
        if isinstance(d, (int, float)):
            gs = float(d)
        if isinstance(e, (int, float)):
            usd = float(e)
        # formulas like =+E4*$D$1 → use E
        if gs is None and usd is not None:
            bancos.append({"label": str(label).strip(), "usd": usd})
        elif gs is not None:
            bancos.append({"label": str(label).strip(), "gs": gs})

    # meses: leer bloques con fecha en col B
    meses: dict[str, dict] = {}
    cur_ym = None
    key_map = {
        "CHEQUES A VENCER": "cheques",
        "SALDO DE CLIENTES": "saldo_clientes",
        "MERCADERIAS A ENTREGAR": "mercaderias",
        "PAGOS BAZZAR": "bazzar",
        "PAGOS DE BAZZAR": "bazzar",
        "PV Y PROG A COBRAR": "pv_prog",
        "PAGO LUISITO": "luisito",
        "PAGO A PROVEEDORES": "proveedores",
        "GASTOS DE DESPACHO": "despacho",
        "PREVISION GASTOS OPERATIVOS": "gastos_op",
        "PRESTAMO BANCARIO": "prestamo",
    }
    for r in range(1, min(ws.max_row, 120) + 1):
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        d = ws.cell(r, 4).value
        if isinstance(b, datetime):
            cur_ym = f"{b.year}-{b.month:02d}"
            meses.setdefault(cur_ym, {})
        if not c or not cur_ym:
            continue
        label = str(c).strip().upper().rstrip()
        for k, field in key_map.items():
            if label.startswith(k):
                if isinstance(d, (int, float)):
                    meses.setdefault(cur_ym, {})[field] = float(d)
                break

    return {"tasa_usd": float(tasa), "bancos": bancos, "meses": meses}


def _parse_corte_from_name(carpeta: Path) -> date:
    # corte-AL-03-08-26 or folder with 03-08
    import re

    m = re.search(r"(\d{2})-(\d{2})-(\d{2,4})", carpeta.name)
    if m:
        dd, mm, yy = m.groups()
        y = int(yy) if len(yy) == 4 else 2000 + int(yy)
        return date(y, int(mm), int(dd))
    return date(2026, 8, 3)


def run(
    entrada: Path,
    salida: Path,
    *,
    persist_local: bool = False,
    supabase: bool = False,
    auto_aprobar_huellas: bool = False,
) -> dict:
    salida.mkdir(parents=True, exist_ok=True)
    csv_dir = salida / "csv"
    csv_dir.mkdir(exist_ok=True)

    clasifs = clasificar_carpeta(entrada)
    (salida / "clasificacion.json").write_text(
        json.dumps([c.to_dict() for c in clasifs], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    cheques_por_mes: dict[str, int] = {}
    saldo_clientes_por_mes: dict[str, int] = {}
    pv_prog_por_mes: dict[str, int] = {}
    aging: dict[str, int] = {}
    reportes = []
    parsed_by_name: dict[str, dict] = {}

    for c in clasifs:
        path = Path(c.path)
        data = parse_por_tipo(c.tipo, path)
        parsed_by_name[c.nombre] = data
        reportes.append(
            {
                "archivo": c.nombre,
                "tipo": c.tipo,
                "confianza": c.confianza,
                "programa_erp": c.programa_erp,
                "totales": data.get("totales", {}),
                "filas": len(data.get("filas") or []),
            }
        )
        filas = data.get("filas") or []
        if filas:
            escribir_csv(filas, csv_dir / f"{path.stem}__{c.tipo}.csv")

        if c.tipo == "cheques_vencer":
            ym = mes_desde_nombre_cheques(c.nombre) or "????"
            cheques_por_mes[ym] = cheques_por_mes.get(ym, 0) + int(
                data.get("totales", {}).get("importe_gs", 0)
            )
        elif c.tipo in ("saldos_resumen", "saldos"):
            saldo_clientes_por_mes["_total_resumen"] = int(
                data.get("totales", {}).get("saldo_gs", 0)
            )
        elif c.tipo == "saldos_detallado":
            aging = data.get("totales", {}).get("aging", {}) or {}
            saldo_clientes_por_mes["_total_detallado"] = int(
                data.get("totales", {}).get("saldo_gs", 0)
            )
        elif c.tipo == "pv_prog":
            pv_prog_por_mes.update(
                {
                    k: int(v)
                    for k, v in (data.get("totales", {}).get("por_mes_cuota") or {}).items()
                }
            )

    ref = entrada / "SF AL 03-08.xlsx"
    manuales = extraer_manuales_referencia(ref)
    # mes corte agosto: valores literales filas 11-22 del ref si existen
    corte = _parse_corte_from_name(entrada)
    ym0 = f"{corte.year}-{corte.month:02d}"
    if openpyxl and ref.exists():
        wb = openpyxl.load_workbook(ref, data_only=False)
        ws = wb["SIT FIN"]
        # agosto implícito: CHEQUES fila 11 vacía; sept empieza 24
        # leer literales 19-22 etc
        ago = manuales.setdefault("meses", {}).setdefault(ym0, {})
        for r, field in [
            (19, "proveedores"),
            (20, "despacho"),
            (21, "gastos_op"),
            (22, "prestamo"),
        ]:
            d = ws.cell(r, 4).value
            e = ws.cell(r, 5).value
            if isinstance(d, (int, float)):
                ago[field] = float(d)
            elif isinstance(e, (int, float)) and field == "proveedores":
                ago[field] = -abs(float(e) * float(manuales["tasa_usd"]))

    # completar meses siguientes desde ref (ya en manuales["meses"])
    # override cheques AUTO siempre
    for ym, v in list(cheques_por_mes.items()):
        if ym.endswith("+"):
            continue
        manuales.setdefault("meses", {}).setdefault(ym, {})["cheques"] = v

    meta = {
        "corte_id": entrada.name,
        "n_archivos": len(clasifs),
        "desconocidos": [c.nombre for c in clasifs if c.tipo == "desconocido"],
    }

    xlsx_out = salida / f"SF_NEXUS_{corte.strftime('%d-%m-%y')}.xlsx"
    generar_sit_fin(
        corte=corte,
        tasa_usd=float(manuales.get("tasa_usd") or 5970.96),
        bancos=manuales.get("bancos") or [
            {"label": "SALDO EN USD. BANCO CONTINENTAL", "usd": 154933.41},
            {"label": "SALDO EN GS. BANCO CONTINENTAL", "gs": 1718236201},
        ],
        cheques_por_mes={k: v for k, v in cheques_por_mes.items() if not k.startswith("_")},
        saldo_clientes_por_mes={
            k: v for k, v in saldo_clientes_por_mes.items() if not k.startswith("_")
        },
        pv_prog_por_mes=pv_prog_por_mes,
        aging=aging,
        manuales=manuales,
        dest=xlsx_out,
        meta=meta,
    )

    # comparación rápida vs referencia
    cmp = {"cheques_nexus": cheques_por_mes, "aging_nexus": aging, "pv_prog": pv_prog_por_mes}
    if openpyxl and ref.exists():
        wb = openpyxl.load_workbook(ref, data_only=False)
        ws = wb["SIT FIN"]
        ref_cheques = {}
        for r in range(1, 70):
            b = ws.cell(r, 2).value
            c = ws.cell(r, 3).value
            d = ws.cell(r, 4).value
            if c and "CHEQUES A VENCER" in str(c).upper() and isinstance(d, (int, float)):
                if hasattr(b, "year"):
                    ref_cheques[f"{b.year}-{b.month:02d}"] = float(d)
        cmp["cheques_ref_excel"] = ref_cheques
        cmp["delta_cheques"] = {
            ym: cheques_por_mes.get(ym, 0) - ref_cheques.get(ym, 0)
            for ym in sorted(set(cheques_por_mes) | set(ref_cheques))
            if not str(ym).endswith("+")
        }

    persist_info: dict | None = None
    if persist_local or supabase:
        cat = CatalogoLocal()
        corte_db = persistir_corte_local(
            catalogo=cat,
            fecha_al=corte,
            tasa_usd=float(manuales.get("tasa_usd") or 5970.96),
            carpeta=str(entrada),
            clasifs=clasifs,
            parsed_by_name=parsed_by_name,
            cheques_por_mes=cheques_por_mes,
            aging=aging,
            pv_prog_por_mes=pv_prog_por_mes,
            manuales=manuales,
            auto_aprobar_huellas=auto_aprobar_huellas,
        )
        persist_info = {
            "batch_id": corte_db["batch_id"],
            "estado": corte_db["estado"],
            "n_variaciones": corte_db["n_variaciones"],
            "staging_counts": corte_db["staging_counts"],
            "catalogo": str(cat.root),
        }
        (salida / "persistencia.json").write_text(
            json.dumps(
                {
                    "batch_id": corte_db["batch_id"],
                    "estado": corte_db["estado"],
                    "n_variaciones": corte_db["n_variaciones"],
                    "variaciones": corte_db["variaciones"],
                    "staging_counts": corte_db["staging_counts"],
                },
                ensure_ascii=False,
                indent=2,
                default=str,
            ),
            encoding="utf-8",
        )
        if supabase:
            sb = persistir_supabase(corte_db)
            persist_info["supabase"] = sb

    resumen = {
        "entrada": str(entrada),
        "salida_xlsx": str(xlsx_out),
        "clasificacion": reportes,
        "meta": meta,
        "comparacion": cmp,
        "persistencia": persist_info,
        "totales_clave": {
            "cheques_por_mes": cheques_por_mes,
            "aging": aging,
            "pv_prog_por_mes": pv_prog_por_mes,
            "saldo_clientes_totales": {
                k: v for k, v in saldo_clientes_por_mes.items() if k.startswith("_")
            },
        },
    }
    (salida / "resumen.json").write_text(
        json.dumps(resumen, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )

    # HTML simple mejorado (presentación)
    html = _html_resumen(resumen, corte)
    (salida / f"SF_NEXUS_{corte.strftime('%d-%m-%y')}.html").write_text(html, encoding="utf-8")
    return resumen


def _html_resumen(resumen: dict, corte: date) -> str:
    cheques = resumen["totales_clave"]["cheques_por_mes"]
    aging = resumen["totales_clave"]["aging"]
    rows_c = "".join(
        f"<tr><td>{k}</td><td class='n'>{v:,}</td></tr>" for k, v in sorted(cheques.items())
    )
    rows_a = "".join(
        f"<tr><td>{k}</td><td class='n'>{v:,}</td></tr>" for k, v in sorted(aging.items())
    )
    rows_cls = "".join(
        f"<tr><td>{r['archivo']}</td><td>{r['tipo']}</td><td>{r['confianza']}</td>"
        f"<td>{r.get('programa_erp') or ''}</td><td>{r['filas']}</td></tr>"
        for r in resumen["clasificacion"]
    )
    return f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="utf-8"/>
<title>SF Nexus · {corte.isoformat()}</title>
<style>
:root {{ --ink:#0f172a; --teal:#0f3d3e; --bg:#f4f7f6; --card:#fff; --auto:#dbeafe; }}
body {{ margin:0; font-family: "Segoe UI", system-ui, sans-serif; background:var(--bg); color:var(--ink); }}
header {{ background:linear-gradient(135deg,#0f3d3e,#1a5c5e); color:#fff; padding:2rem 2.5rem; }}
header h1 {{ margin:0; font-weight:600; letter-spacing:-.02em; }}
header p {{ margin:.4rem 0 0; opacity:.85; }}
main {{ max-width:1100px; margin:0 auto; padding:1.5rem; display:grid; gap:1.25rem; }}
section {{ background:var(--card); border-radius:12px; padding:1.25rem 1.5rem; box-shadow:0 1px 3px rgba(0,0,0,.06); }}
h2 {{ margin:0 0 1rem; font-size:1.05rem; color:var(--teal); }}
table {{ width:100%; border-collapse:collapse; font-size:.92rem; }}
th,td {{ text-align:left; padding:.45rem .5rem; border-bottom:1px solid #e5e7eb; }}
td.n {{ text-align:right; font-variant-numeric:tabular-nums; }}
.badge {{ display:inline-block; background:var(--auto); color:#1e3a8a; padding:.15rem .5rem; border-radius:6px; font-size:.75rem; }}
</style></head><body>
<header>
  <h1>Situación financiera Nexus</h1>
  <p>Corte {corte.strftime('%d/%m/%Y')} · TXT clasificados · linaje AUTO · objetivo 100% presentación</p>
</header>
<main>
<section>
  <h2>Clasificación de TXT <span class="badge">huellas ERP</span></h2>
  <table><thead><tr><th>Archivo</th><th>Tipo</th><th>Conf.</th><th>Programa</th><th>Filas</th></tr></thead>
  <tbody>{rows_cls}</tbody></table>
</section>
<section>
  <h2>Cheques a vencer (AUTO)</h2>
  <table><thead><tr><th>Mes</th><th>Importe Gs</th></tr></thead><tbody>{rows_c}</tbody></table>
</section>
<section>
  <h2>Aging saldo clientes (AUTO)</h2>
  <table><thead><tr><th>Bucket</th><th>Saldo Gs</th></tr></thead><tbody>{rows_a}</tbody></table>
</section>
<section>
  <p>Excel: <code>{resumen.get('salida_xlsx','')}</code></p>
</section>
</main></body></html>"""


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--entrada", type=Path, default=DEFAULT_IN)
    ap.add_argument("--salida", type=Path, default=DEFAULT_OUT)
    ap.add_argument(
        "--persist-local",
        action="store_true",
        help="Persiste T01–T12 en data/catalogo_local + check variaciones",
    )
    ap.add_argument(
        "--supabase",
        action="store_true",
        help="También inserta en Postgres LAB (DATABASE_URL / SUPABASE_DB_URL)",
    )
    ap.add_argument(
        "--auto-aprobar-huellas",
        action="store_true",
        help="Aprueba huellas vistas en esta corrida (solo LAB / seed)",
    )
    args = ap.parse_args()
    resumen = run(
        args.entrada.resolve(),
        args.salida.resolve(),
        persist_local=args.persist_local or args.supabase,
        supabase=args.supabase,
        auto_aprobar_huellas=args.auto_aprobar_huellas,
    )
    print("OK", resumen["salida_xlsx"])
    print("Clasificados:", len(resumen["clasificacion"]))
    print("Desconocidos:", resumen["meta"]["desconocidos"])
    print("Cheques/mes:", resumen["totales_clave"]["cheques_por_mes"])
    if "delta_cheques" in resumen.get("comparacion", {}):
        print("Delta vs Excel ref:", resumen["comparacion"]["delta_cheques"])
    if resumen.get("persistencia"):
        p = resumen["persistencia"]
        print(
            f"Persistencia: estado={p['estado']} variaciones={p['n_variaciones']} "
            f"batch={p['batch_id']}"
        )
        if p.get("supabase"):
            print("Supabase:", p["supabase"])


if __name__ == "__main__":
    main()
