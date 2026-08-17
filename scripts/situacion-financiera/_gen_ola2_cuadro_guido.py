# -*- coding: utf-8 -*-
"""Ola 2/4 Guido — Cuadro + Detalle → Sit Fin proyección por mes.

Cablea **todas** las columnas de proyección del pivot Cuadro (no solo ago-26)
y reemplaza el árbol molecular de cada mes con cuotas filtradas desde Detalle
(FILA_CUADRO × COLUMNA_CUADRO), excluyendo SALEMMA/DIFICIL/BAZZAR del
SALDO DE CLIENTES.

Reglas (G4/G7/G8 · Comentario General + observación Guido 2026-08-13):
  - SALDO DE CLIENTES {mes} = fila OK × col mes
  - MERCADERÍAS {mes} = fila A ENTREGAR × col mes
  - PAGO LUISITO {mes} = fila LUISITO × col mes
  - VENCIDOS 30/60 (corte ago) = OK + LUISITO × jul/jun-26

Fuente canon: Excel Guido 08 · hojas Cuadro + Detalle
"""
from __future__ import annotations

import json
import sys
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
SF = ROOT / "src/lib/situacion-financiera"
MOL = SF / "molecular-al-0308.json"
OUT = SF / "ola2-cuadro-0308.json"

EXCEL_GUIDO = None  # resuelto en main()

TASA = 5970.96

FILAS_CUADRO = ("A ENTREGAR", "OK", "LUISITO", "DIFICIL", "SALEMMA")

# Columna Cuadro → mol_key YYYY-MM
CUADRO_COL_TO_YM: dict[str, str] = {
    "ago-26": "2026-08",
    "sept-26": "2026-09",
    "oct-26": "2026-10",
    "nov-26": "2026-11",
    "dic-26": "2026-12",
    "ene-27": "2026-01",
}

MOL_PREFIX_BY_FILA: dict[str, str] = {
    "OK": "clientes",
    "A ENTREGAR": "mercaderia",
    "LUISITO": "luisito",
}


def _excel() -> Path:
    global EXCEL_GUIDO
    if EXCEL_GUIDO is None:
        from _guido_excel_path import resolver_excel_guido_08

        EXCEL_GUIDO = resolver_excel_guido_08()
    return EXCEL_GUIDO


def _num(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _usd(gs: float) -> float:
    return round(gs / TASA, 2)


def extraer_cuadro_excel(path: Path) -> dict:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Cuadro"]
    rows = list(ws.iter_rows(min_row=1, max_row=40, values_only=True))
    wb.close()

    header = None
    header_idx = None
    for i, row in enumerate(rows):
        if row and row[0] and "Etiquetas de fila" in str(row[0]):
            header = [str(c).strip().lower() if c is not None else "" for c in row]
            header_idx = i
            break
    if not header:
        raise RuntimeError("No se encontró fila encabezado Cuadro (Etiquetas de fila)")

    col_map = {name: idx for idx, name in enumerate(header) if name}

    def col_mes(abrev: str) -> int:
        key = abrev.lower().strip()
        if key not in col_map:
            raise KeyError(f"Columna {abrev!r} no en Cuadro: {list(col_map.keys())}")
        return col_map[key]

    data: dict[str, dict[str, float]] = {}
    for row in rows[header_idx + 1 :]:
        if not row or row[0] is None:
            continue
        fila = str(row[0]).strip().upper()
        if fila not in FILAS_CUADRO:
            continue
        data[fila] = {}
        for name, idx in col_map.items():
            if name in ("etiquetas de fila", "total general"):
                continue
            if idx < len(row):
                data[fila][name] = _num(row[idx])

    def cel(fila: str, mes: str) -> float:
        return data.get(fila, {}).get(mes.lower(), 0.0)

    metricas: dict[str, dict] = {}

    # Proyección por mes (Ola 4 — Guido 13/08)
    for col, ym in CUADRO_COL_TO_YM.items():
        gs_ok = cel("OK", col)
        if gs_ok or col == "ago-26":
            metricas[f"clientes:{ym}"] = {
                "gs": gs_ok,
                "regla": f'G4 · OK × {col}',
                "fila_cuadro": "OK",
                "columna": col,
            }
        gs_ent = cel("A ENTREGAR", col)
        if gs_ent or col == "ago-26":
            metricas[f"mercaderia:{ym}"] = {
                "gs": gs_ent,
                "regla": f'G7 · A ENTREGAR × {col}',
                "fila_cuadro": "A ENTREGAR",
                "columna": col,
            }
        gs_lui = cel("LUISITO", col)
        if gs_lui or col == "ago-26":
            metricas[f"luisito:{ym}"] = {
                "gs": gs_lui,
                "regla": f'G8 · LUISITO × {col}',
                "fila_cuadro": "LUISITO",
                "columna": col,
            }
        gs_dif = cel("DIFICIL", col) + cel("SALEMMA", col)
        if gs_dif:
            metricas[f"dificil:{ym}"] = {
                "gs": gs_dif,
                "regla": f'DIF.COBRO · {{DIFICIL,SALEMMA}} × {col}',
                "fila_cuadro": "DIFICIL+SALEMMA",
                "columna": col,
            }

    # Aging corte agosto (Ola 2 original)
    jul = "jul-26"
    jun = "jun-26"
    metricas["aging:v30"] = {
        "gs": cel("OK", jul) + cel("LUISITO", jul),
        "regla": "venc30 · {OK,LUISITO} × jul-26",
        "fila_cuadro": "OK+LUISITO",
        "columna": jul,
    }
    metricas["aging:v60"] = {
        "gs": cel("OK", jun) + cel("LUISITO", jun),
        "regla": "venc60 · {OK,LUISITO} × jun-26",
        "fila_cuadro": "OK+LUISITO",
        "columna": jun,
    }

    return {
        "fuente": str(path),
        "hoja": "Cuadro",
        "columnas": {k: v for k, v in col_map.items() if k not in ("etiquetas de fila",)},
        "filas_raw": data,
        "metricas": metricas,
    }


def _child_node(
    nid: str,
    label: str,
    gs: float,
    meta: str,
    fuente: str,
    children: list[dict] | None = None,
) -> dict:
    n: dict = {
        "id": nid,
        "label": label,
        "gs": gs,
        "usd": _usd(gs),
        "meta": meta,
        "fuente": fuente,
    }
    if children:
        n["children"] = children
    return n


def build_detalle_children(path: Path, fila_cuadro: str, col_cuadro: str) -> list[dict]:
    """Árbol cliente → factura desde Detalle · solo filas del bucket Cuadro."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Detalle"]
    by_cli: dict[str, dict] = defaultdict(lambda: {"nombre": "", "facturas": []})
    col = col_cuadro.lower().strip()
    fila = fila_cuadro.upper().strip()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 14:
            continue
        if str(row[7] or "").strip().upper() != fila:
            continue
        if str(row[11] or "").strip().lower() != col:
            continue
        cuota = _num(row[13])
        if cuota == 0:
            continue
        cod = str(row[2] or "").strip()
        nom = str(row[1] or "").strip()
        nro = str(row[0] or "").strip()
        by_cli[cod]["nombre"] = nom
        by_cli[cod]["facturas"].append(
            {
                "nro": nro,
                "cuota": cuota,
                "venc": row[10],
                "tipo": str(row[6] or "").strip(),
            }
        )
    wb.close()

    nodes: list[dict] = []
    for cod, info in sorted(by_cli.items(), key=lambda x: -sum(f["cuota"] for f in x[1]["facturas"])):
        facs = info["facturas"]
        gs_cli = sum(f["cuota"] for f in facs)
        children = [
            _child_node(
                f"det-{col}-{fila}-{cod}-{i}-{f['nro']}",
                f"Factura {f['nro']}",
                f["cuota"],
                f"cuota · venc {f['venc']} · {f['tipo']}",
                "Detalle Excel Guido",
            )
            for i, f in enumerate(sorted(facs, key=lambda x: -x["cuota"]))
        ]
        nodes.append(
            _child_node(
                f"det-{col}-{fila}-cli-{cod}",
                f"{cod} · {info['nombre'] or 'sin nombre'}",
                gs_cli,
                f"{len(facs)} cuota(s) · {fila} × {col}",
                "Detalle Excel Guido",
                children=children,
            )
        )
    return nodes


def parchear_molecular(cuadro: dict, excel: Path) -> dict[str, float]:
    mol = json.loads(MOL.read_text(encoding="utf-8"))
    aplicados: dict[str, float] = {}
    fuente = f"Ola2/4 Cuadro+Detalle Guido · {cuadro['fuente']}"

    for mol_key, info in cuadro["metricas"].items():
        gs = float(info["gs"])
        aplicados[mol_key] = gs
        prev = mol.get(mol_key, {})
        ola_tag = "Ola 4" if mol_key.startswith(("clientes:2026-09", "mercaderia:2026-09")) else "Ola 2/4"

        children = prev.get("children")
        fila = info.get("fila_cuadro", "")
        col = info.get("columna", "")
        if fila in MOL_PREFIX_BY_FILA and col in CUADRO_COL_TO_YM:
            children = build_detalle_children(excel, fila, col)
            # Verificar 1323 SALEMMA no aparece en clientes OK
            if mol_key == "clientes:2026-09":
                for ch in children:
                    if "1323" in ch.get("label", "") or "SALEMMA" in ch.get("label", "").upper():
                        raise RuntimeError("SALEMMA en árbol OK sept — revisar Detalle")

        node: dict = {
            **prev,
            "id": prev.get("id") or mol_key.replace(":", "-"),
            "label": prev.get("label") or mol_key,
            "gs": gs,
            "usd": _usd(gs),
            "meta": (
                f"{ola_tag} · {info['regla']} · {info['fila_cuadro']} × {info['columna']} "
                f"(antes={prev.get('gs')})"
            ),
            "fuente": fuente,
            "ola2": True,
            "ola4": bool(col != "ago-26" or mol_key.endswith("2026-09")),
        }
        if children is not None:
            node["children"] = children
        mol[mol_key] = node

        if mol_key == "luisito:2026-08" and gs:
            mol["luisito:cuadro"] = {
                **mol.get("luisito:cuadro", {}),
                "gs": gs,
                "usd": _usd(gs),
                "meta": f"Ola 2/4 · celda mes LUISITO · {col}",
                "fuente": fuente,
                "ola2": True,
            }

    if "pv:2026-08" in mol and mol["pv:2026-08"].get("gs") == mol.get("mercaderia:2026-08", {}).get("gs"):
        mol["mercaderia:2026-08"]["meta"] = (
            str(mol["mercaderia:2026-08"].get("meta", "")) + " · separado de pv:2026-08"
        )

    MOL.write_text(json.dumps(mol, ensure_ascii=False, indent=2), encoding="utf-8")
    return aplicados


def main() -> None:
    excel = _excel()
    if not excel.exists():
        print("FAIL: Excel Guido no encontrado:", excel, file=sys.stderr)
        raise SystemExit(1)
    if not MOL.exists():
        print("FAIL: molecular ausente — correr _gen_molecular_al.py primero", file=sys.stderr)
        raise SystemExit(1)

    cuadro = extraer_cuadro_excel(excel)
    aplicados = parchear_molecular(cuadro, excel)

    payload = {
        "corte": "AL-03-08-26",
        "ola": "2+4",
        "actualizado": "2026-08-17",
        "excel_canon": str(excel),
        "pipeline": "TXT → Detalle → TIPO COBRO → Cuadro → Sit Fin verdes",
        "metricas": cuadro["metricas"],
        "aplicados_molecular": aplicados,
        "nota": "Ola 4 · proyección todos los meses Cuadro + árbol Detalle (Guido 13/08)",
        "observacion_guido": "SALDO CLIENTES sept solo OK — excluir SALEMMA/DIFICIL/BAZZAR",
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    print("OLA2/4 OK", OUT)
    for k in sorted(aplicados):
        if k.startswith(("clientes:", "mercaderia:", "luisito:")):
            print(f"  {k}: {aplicados[k]:,.0f} Gs")


if __name__ == "__main__":
    main()
