import sys
from pathlib import Path

from openpyxl import load_workbook

p = Path(r"C:\Users\hecto\Nexus_Core\csv's\programado\casos francis.xlsx")
wb = load_workbook(p, read_only=True, data_only=True)
ws = wb.active
for i, row in enumerate(ws.iter_rows(max_row=30, max_col=20, values_only=True)):
    if any(c is not None for c in row):
        print(i + 1, row)
